import os
import requests
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
from queue import Queue

def download_image(place_name, image_url, output_folder):
    image_filename = f"{place_name}.jpg"
    image_path = os.path.join(output_folder, image_filename)
    
    response = requests.get(image_url)
    if response.status_code == 200:
        with open(image_path, 'wb') as f:
            f.write(response.content)
        print(f"Downloaded {image_filename}")
    else:
        print(f"Failed to download {image_filename}")

def download_images_from_excel(excel_file_path, output_folder):
    wb = load_workbook(excel_file_path)
    sheet = wb.active

    headers = [cell.value.lower() for cell in sheet[1]]
    place_index = headers.index('place') + 1
    image_url_index = headers.index('image_url') + 1

    # Create a queue to maintain the order of download tasks
    task_queue = Queue()

    # Fill the queue with download tasks
    for row in sheet.iter_rows(min_row=2, values_only=True):
        place_name, image_url = row[place_index - 1], row[image_url_index - 1]
        task_queue.put((place_name, image_url))

    # Use ThreadPoolExecutor to execute download tasks
    with ThreadPoolExecutor(max_workers=5) as executor:
        while not task_queue.empty():
            place_name, image_url = task_queue.get()
            executor.submit(download_image, place_name, image_url, output_folder)

if __name__ == "__main__":
    excel_file_path = r"C:\Users\mercy\Downloads\place\places.xlsx"
    output_folder = r"C:\Users\mercy\Downloads\place\images"
    os.makedirs(output_folder, exist_ok=True)
    download_images_from_excel(excel_file_path, output_folder)
